import os
from datetime import datetime
from faker import Faker
import openpyxl
from openpyxl.styles import Font
fake= Faker("en_IN")  # Initialize Faker with Hindi locale
from models.employee import Employee
employees= []

for _ in  range(1,10):
    
    employee=Employee(
        fake.random_int(min=1000, max=9999),
        fake.name(),
        fake.job(),
        fake.random_int(min=30000, max=120000),
        fake.email(),
        fake.phone_number()
       
    )
    employees.append(employee)
sorted_employees = sorted(employees, key=lambda x: x.get_name(), reverse=False)

dirPath="E:\\MTSTraningmay2025\\day1\\reports"
if not os.path.exists(dirPath):
    os.makedirs(dirPath)

wb= openpyxl.Workbook()
ws1 = wb.create_sheet("May-2025")
#ws = wb.active  
#ws.title = "Employee Report"
# Write the header  
ws1.append(["Id", "Name", "Designation", "Salary", "Email", "Phone"])
# Write the employee data   
for employee in sorted_employees:  
    ws1.append([
        employee.get_id(),
        employee.get_name(),
        employee.get_designation(),
        employee.get_salary(),
        employee.email,
        employee.get_contact_no()
    ])
for row in ws1.iter_rows():
    for cell in row:
        cell.font =  Font(name='Arial', size=24, bold=True if cell.row == 1 else False)
filePath=os.path.join(dirPath, f"employee_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
wb.save(filePath)
wb.close()